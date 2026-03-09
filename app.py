# -*- coding: utf-8 -*-
"""
Sistema Gestión de Usuarios - PUREM Industrial
"""

import streamlit as st
import pandas as pd
import firebase_admin
from firebase_admin import credentials, firestore
import datetime
import time
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------- INIT FIRESTORE ----------------
if "db" not in st.session_state:
    if not firebase_admin._apps:
        firebase_config = dict(st.secrets["firebase"])
        cred = credentials.Certificate(firebase_config)
        firebase_admin.initialize_app(cred)
    st.session_state.db = firestore.client()

db = st.session_state.db

# ---------------- CONFIG ----------------
st.set_page_config(
    page_title="Gestión de Usuarios",
    page_icon="logoo.png",
    layout="wide"
)

# ---------------- ESTILOS ----------------
st.markdown("""
<style>
html, body, .main {
    background-color: #0E1117;
    color: white;
    font-family: 'Montserrat', sans-serif;
}
/* MEJORA LOGIN CELULAR */
@media (max-width: 800px) {
    .login-banner {
        display: none !important;
    }
    .stColumn {
        width: 100% !important;
    }
}
.card {
    background-color: #1b1f2a;
    padding: 1.2em;
    border-radius: 12px;
    margin-bottom: 1em;
    transition: 0.3s ease;
}
.card:hover {
    transform: translateY(-5px);
    box-shadow: 0 8px 20px rgba(0,255,153,0.2);
}
.badge-tecnico {
    background-color: #009966;
    padding: 4px 10px;
    border-radius: 6px;
    font-size: 0.8rem;
}
.badge-supervisor {
    background-color: #0055aa;
    padding: 4px 10px;
    border-radius: 6px;
    font-size: 0.8rem;
}
.badge-admin {
    background-color: #003366;
    padding: 4px 10px;
    border-radius: 6px;
    font-size: 0.8rem;
}
.stButton>button {
    background: linear-gradient(135deg, #00ff99 0%, #009966 100%);
    border-radius: 8px;
    font-weight: 600;
    transition: 0.3s ease;
}
.stButton>button:hover {
    transform: translateY(-3px);
    box-shadow: 0 8px 20px rgba(0,255,153,0.4);
}
.footer {
    position: fixed;
    bottom: 10px;
    right: 20px;
    color: rgba(255,255,255,0.5);
    font-size: 0.9rem;
}
.login-footer {
    margin-top: 15px;
    color: rgba(255,255,255,0.4);
    font-size: 0.9rem;
}
.group-heading {
    margin-top: 1.5rem;
    color: #00ff99;
    font-size: 1.2rem;
    font-weight: bold;
    border-bottom: 1px solid #333;
    padding-bottom: 5px;
}
.week-heading {
    margin-top: 0.8rem;
    color: #00ccff;
    font-weight: 600;
}
.small-note {
    color: rgba(255,255,255,0.7);
    font-size: 0.85rem;
}
</style>
""", unsafe_allow_html=True)

# ---------------- SESSION ----------------
if "auth" not in st.session_state:
    st.session_state.auth = False
if "user" not in st.session_state:
    st.session_state.user = None
if "log_df" not in st.session_state:
    st.session_state.log_df = pd.DataFrame()
if "excel_buffer" not in st.session_state:
    st.session_state.excel_buffer = None

# ---------------- FUNCIONES ----------------

def generar_excel_usuarios(usuarios):
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios Registrados"
    ws.merge_cells("A1:H1")
    ws["A1"] = "REPORTE OFICIAL DE USUARIOS REGISTRADOS - PUREM"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])
    headers = ["MX", "Nombre", "Unidad", "Business Unit", "Emp No", "Role", "Área", "Última Alta"]
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=3, column=col)
        cell.fill = PatternFill(start_color="00331A", end_color="00331A", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for u in usuarios:
        last_login = u.get("last_login")
        # last_login puede venir como timestamp de Firestore (aware) o string
        if isinstance(last_login, datetime.datetime):
            # normalizamos para mostrar
            last_login_str = last_login.strftime("%d-%m-%Y %H:%M")
        else:
            last_login_str = str(last_login) if last_login else "-"
        ws.append([
            u.get("mx", "-"),
            u.get("nombre", "-"),
            u.get("unidad", "-"),
            u.get("business_unit", "-"),
            u.get("emp_no", "-"),
            u.get("role", "-"),
            u.get("area", "-"),
            last_login_str
        ])
    for col in ws.columns:
        max_length = max([len(str(cell.value)) for cell in col if cell.value] + [10])
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 4
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def actualizar_excel_general():
    docs = db.collection("empleados").stream()
    usuarios = []
    for d in docs:
        u = d.to_dict()
        u["mx"] = d.id
        usuarios.append(u)
    excel_bytes = generar_excel_usuarios(usuarios)
    st.session_state.excel_buffer = excel_bytes
    return usuarios, excel_bytes

def parse_fecha_de_registro(h):
    ts = h.get("fecha_ts")
    if isinstance(ts, datetime.datetime):
        # normalize offset-aware to naive for consistent comparisons
        return ts.replace(tzinfo=None)
    f = h.get("fecha")
    if f and isinstance(f, str):
        for fmt in ("%d-%m-%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.datetime.strptime(f, fmt)
            except:
                continue
    return None

def agrupar_historial_por_mes_semana(hist_docs):
    groups = {}
    items = []
    for doc in hist_docs:
        data = doc.to_dict()
        dt = parse_fecha_de_registro(data)
        items.append((doc.id, data, dt))
    # ordenar por fecha desc (dt puede ser None)
    items_sorted = sorted(items, key=lambda x: x[2] or datetime.datetime(1970,1,1), reverse=True)
    for doc_id, data, dt in items_sorted:
        if dt:
            month_key = f"{dt.year}-{dt.month:02d}"
            month_label = f"{dt.strftime('%B').capitalize()} {dt.year}"
            week_no = dt.isocalendar()[1]
        else:
            month_key, month_label, week_no = "SinFecha", "Sin Fecha", 0
        groups.setdefault(month_key, {"label": month_label, "weeks": {}})
        groups[month_key]["weeks"].setdefault(week_no, [])
        groups[month_key]["weeks"][week_no].append((doc_id, data, dt))
    return groups

# ---------------- LOGIN ----------------
if not st.session_state.auth:
    col1, col2 = st.columns([2,2])
    with col1:
        st.markdown("""
        <div class="login-banner" style="background: linear-gradient(135deg, #001a0d 0%, #00331a 100%);
        height:100vh;padding:8% 5%;display:flex;flex-direction:column;justify-content:center;">
        <h1 style="font-size:4rem;">Purem by Eberspächer<br>
        <span style="color:#00ff99;">Mantenimiento Industrial</span></h1>
        <p style="border-left:3px solid #00ff99;padding-left:20px;">
        Plataforma para Gestión de Usuarios de Mantenimiento.</p>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown("## INICIAR SESIÓN")
        mx_input = st.text_input("ID de Usuario (MX)")
        password_input = st.text_input("Contraseña", type="password")
        if st.button("Acceder"):
            if not mx_input or not password_input:
                st.error("Ingrese sus datos")
            else:
                doc = db.collection("empleados").document(mx_input.upper()).get()
                if not doc.exists or doc.to_dict().get("password") != password_input:
                    st.error("Usuario o contraseña incorrecta")
                else:
                    user_data = doc.to_dict()
                    if user_data.get("role") != "admin":
                        st.error("Solo los administradores tienen acceso")
                    else:
                        st.session_state.auth = True
                        st.session_state.user = user_data
                        # precargar excel general al iniciar sesion
                        try:
                            actualizar_excel_general()
                        except:
                            st.session_state.excel_buffer = None
                        st.rerun()
        st.markdown("""
            <p style="font-size: 0.9rem; color: rgba(255,255,255,0.4); text-align: center; margin-top: 60px;">
                SISTEMA PRIVADO DE USO EXCLUSIVO.<br>
                CUALQUIER INTENTO DE ACCESO NO AUTORIZADO SERÁ MONITOREADO.
            </p>
            <p style="font-size: 1.0rem; color: rgba(255,255,255,0.3); text-align: center; margin-top: 20px;">
                PUREM BY EBERSPÄCHER - RAMOS ARIZPE<br>
                DEVELOPED BY: JUAN RODRIGO GUZMÁN MARTÍNEZ
            </p>
        """, unsafe_allow_html=True)
        #st.markdown('<div class="login-footer">Developed by: Juan Rodrigo Guzmán Martínez</div>', unsafe_allow_html=True)

# ---------------- PANEL ADMIN ----------------
if st.session_state.auth:
    with st.sidebar:
        st.image("logoo.png", width=120)
        st.markdown(f"### {st.session_state.user['nombre']}")
        if st.button("Cerrar sesión"):
            st.session_state.auth = False
            st.session_state.user = None
            st.session_state.log_df = pd.DataFrame()
            st.session_state.excel_buffer = None
            st.rerun()

    st.title("👥 Panel de Administración")
    tab1, tab_manual, tab2, tab3 = st.tabs(["📤 Alta de usuarios", "➕ Alta Manual", "📋 Usuarios registrados", "📜 Historial de altas"])

    # ---------------- TAB 1: ALTA EXCEL ----------------
    with tab1:
        st.subheader("Altas de Usuarios de Forma Masiva")
        uploaded_file = st.file_uploader("Subir archivo Excel", type=["xlsx"])
        if uploaded_file:
            try:
                df = pd.read_excel(uploaded_file)
            except Exception:
                st.error("Error leyendo el archivo. Asegúrate de que sea un .xlsx válido.")
                df = None
            if df is not None:
                st.info(f"{len(df)} registros detectados")
                if st.button("Procesar archivo"):
                    progress = st.progress(0)
                    resultados = []
                    total = len(df)
                    for i, row in df.iterrows():
                        mx_id = str(row["mx"]).upper()
                        doc_ref = db.collection("empleados").document(mx_id)
                        doc = doc_ref.get()
                        estado = "NO MODIFICADO"
                        if not doc.exists:
                            now = datetime.datetime.utcnow()
                            emp_no_val = int(row["emp_no"]) if pd.notna(row.get("emp_no")) else 0
                            doc_ref.set({
                                "mx": mx_id,
                                "nombre": row.get("nombre", ""),
                                "unidad": row.get("unidad", ""),
                                "business_unit": row.get("business_unit", ""),
                                "emp_no": emp_no_val,
                                "password": str(row.get("password", "")),
                                "role": row.get("role", "tecnico"),
                                "area": row.get("area", "General"),
                                "last_login": now,
                                "mantener_sesion": False
                            })
                            estado = "REGISTRADO"
                        resultados.append({
                            "MX": mx_id,
                            "Nombre": row.get("nombre", ""),
                            "Estado": estado,
                            "Procesado por": st.session_state.user["nombre"],
                            "Fecha": datetime.datetime.utcnow().strftime("%d-%m-%Y %H:%M")
                        })
                        progress.progress((i+1)/total)
                        time.sleep(0.02)
                    st.session_state.log_df = pd.DataFrame(resultados)

                    # Actualizar Excel general con todos los usuarios (se guarda en session_state)
                    usuarios, excel_bytes = actualizar_excel_general()

                    # Guardar solo metadata en historial (sin excel_content)
                    try:
                        db.collection("historial_altas").add({
                            "admin": st.session_state.user["nombre"],
                            "fecha": datetime.datetime.utcnow().strftime("%d-%m-%Y %H:%M"),
                            "fecha_ts": datetime.datetime.utcnow(),
                            "registros": len(resultados),
                            "tipo": "alta_masiva"
                        })
                    except Exception as e:
                        st.warning(f"Advertencia: no se pudo registrar el historial: {e}")

                    st.success("Proceso completado y Excel actualizado con todos los usuarios")

        if not st.session_state.log_df.empty:
            st.markdown("### 📊 Resultado de carga")
            st.dataframe(st.session_state.log_df, use_container_width=True)
            if st.session_state.excel_buffer:
                st.download_button(
                    "📥 Descargar reporte actualizado de todos los usuarios",
                    st.session_state.excel_buffer,
                    file_name="reporte_usuarios_actualizado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            if st.button("Limpiar estado"):
                st.session_state.log_df = pd.DataFrame()
                st.session_state.excel_buffer = None
                st.rerun()

    # ---------------- TAB MANUAL ----------------
    with tab_manual:
        st.subheader("Altas de Usuarios de Forma Manual")
        with st.form("manual_form", clear_on_submit=True):
            col_m1, col_m2 = st.columns(2)
            with col_m1:
                m_mx = st.text_input("MX")
                m_nombre = st.text_input("Nombre")
                m_unidad = st.text_input("Unidad")
                m_bu = st.text_input("Business Unit")
                m_emp = st.number_input("Emp No", step=1)
            with col_m2:
                m_pass = st.text_input("Password", type="password")
                m_role = st.selectbox("Role", ["tecnico", "supervisor", "admin"])
                st.text_input("Área", value="General", disabled=True)
                st.checkbox("Mantener Sesión", value=False, disabled=True)
                st.text_input("Last Login", value="None", disabled=True)

            c_btn1, c_btn2 = st.columns([1,4])
            with c_btn1:
                confirmar = st.form_submit_button("Confirmar Alta")
            with c_btn2:
                cancelar = st.form_submit_button("Cancelar")

            if confirmar:
                if not m_mx or not m_nombre or not m_pass:
                    st.error("Llena los campos obligatorios (MX, Nombre, Password)")
                else:
                    mx_upper = m_mx.strip().upper()
                    now = datetime.datetime.utcnow()
                    doc_ref = db.collection("empleados").document(mx_upper)
                    if doc_ref.get().exists:
                        st.error("El usuario ya existe.")
                    else:
                        try:
                            doc_ref.set({
                                "mx": mx_upper,
                                "nombre": m_nombre.strip(),
                                "unidad": m_unidad.strip(),
                                "business_unit": m_bu.strip(),
                                "emp_no": int(m_emp),
                                "password": m_pass,
                                "role": m_role,
                                "area": "General",
                                "last_login": now,
                                "mantener_sesion": False
                            })
                        except Exception as e:
                            st.error(f"Error al guardar usuario: {e}")
                            st.rerun()

                        # Actualizar Excel general con todos los usuarios (en session)
                        usuarios, excel_bytes = actualizar_excel_general()

                        # Guardar solo metadata en historial (sin excel_content)
                        try:
                            db.collection("historial_altas").add({
                                "admin": st.session_state.user["nombre"],
                                "fecha": now.strftime("%d-%m-%Y %H:%M"),
                                "fecha_ts": now,
                                "target_mx": mx_upper,
                                "target_nombre": m_nombre.strip(),
                                "tipo": "alta_manual"
                            })
                        except Exception as e:
                            st.warning(f"Advertencia: no se pudo registrar el historial: {e}")

                        st.success(f"Usuario {mx_upper} registrado.")
                        st.session_state.excel_buffer = excel_bytes
                        time.sleep(1)
                        st.rerun()

            if cancelar:
                st.experimental_rerun()

    # ---------------- TAB 2: REGISTRADOS ----------------
    with tab2:
        st.subheader("Usuarios Registrados en el Sistema")
        docs = db.collection("empleados").stream()
        usuarios = [{"mx": d.id, **d.to_dict()} for d in docs]
        
        total_tecnicos = len([u for u in usuarios if u.get("role") == "tecnico"])
        total_supervisores = len([u for u in usuarios if u.get("role") == "supervisor"])
        total_admins = len([u for u in usuarios if u.get("role") == "admin"])

        colA, colB, colC = st.columns(3)
        colA.markdown(f"""<div style="background-color:#001f3f;padding:20px;border-radius:10px;text-align:center;"><h3 style="color:#00ff99;">🛠 Técnicos</h3><p style="font-size:2rem;font-weight:bold;">{total_tecnicos}</p></div>""", unsafe_allow_html=True)
        colB.markdown(f"""<div style="background-color:#001f3f;padding:20px;border-radius:10px;text-align:center;"><h3 style="color:#00ff99;">🧑‍💼 Supervisores</h3><p style="font-size:2rem;font-weight:bold;">{total_supervisores}</p></div>""", unsafe_allow_html=True)
        colC.markdown(f"""<div style="background-color:#001f3f;padding:20px;border-radius:10px;text-align:center;"><h3 style="color:#00ff99;">👑 Admins</h3><p style="font-size:2rem;font-weight:bold;">{total_admins}</p></div>""", unsafe_allow_html=True)

        search = st.text_input("🔎 Buscar por MX o Nombre")
        if search:
            usuarios = [u for u in usuarios if search.lower() in u["mx"].lower() or search.lower() in u["nombre"].lower()]

        tecnicos = [u for u in usuarios if u.get("role") == "tecnico"]
        supervisores = [u for u in usuarios if u.get("role") == "supervisor"]

        tab_tec, tab_sup = st.tabs(["🛠 Técnicos", "🧑‍💼 Supervisores"])

        def mostrar_usuario(u):
            badge = "badge-tecnico" if u.get("role") == "tecnico" else "badge-supervisor" if u.get("role") == "supervisor" else "badge-admin"
            confirm_key = f"confirm_delete_{u['mx']}"
            last_login_display = u.get("last_login")
            if isinstance(last_login_display, datetime.datetime):
                last_login_display = last_login_display.strftime("%d-%m-%Y %H:%M")
            st.markdown(f"""<div class="card"><h4 style="color:#00ff99;">{u['nombre']} ({u['mx']}) <span class="{badge}">{u.get('role')}</span></h4><p><b>Unidad:</b> {u.get('unidad','-')} | <b>Área:</b> {u.get('area','-')} | <b>Última Alta:</b> {last_login_display}</p></div>""", unsafe_allow_html=True)
            if not st.session_state.get(confirm_key, False):
                if st.button("⋮ Opciones", key=f"opt_{u['mx']}"):
                    st.session_state[confirm_key] = True
                    st.rerun()
            if st.session_state.get(confirm_key, False):
                st.warning("¿Eliminar definitivamente?")
                cA, cB = st.columns(2)
                if cA.button("Confirmar", key=f"yes_{u['mx']}"):
                    db.collection("empleados").document(u["mx"]).delete()
                    del st.session_state[confirm_key]
                    st.rerun()
                if cB.button("Cancelar", key=f"no_{u['mx']}"):
                    del st.session_state[confirm_key]
                    st.rerun()

        with tab_tec:
            for u in tecnicos: mostrar_usuario(u)
        with tab_sup:
            for u in supervisores: mostrar_usuario(u)

    # ---------------- TAB 3: HISTORIAL ----------------
    with tab3:
        st.subheader("📜 Historial de altas realizadas")
        # Botón único para descargar Excel global con todos los usuarios
        if st.button("📥 Descargar Excel (Todos los usuarios)"):
            try:
                usuarios, excel_bytes = actualizar_excel_general()
                st.download_button(
                    "📥 Descargar archivo generado",
                    excel_bytes,
                    file_name=f"reporte_usuarios_{datetime.datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_global_hist"
                )
            except Exception as e:
                st.error(f"No se pudo generar el Excel: {e}")

        # Mostrar fichas agrupadas (más reciente -> más viejo)
        hist_docs = list(db.collection("historial_altas").stream())
        if not hist_docs:
            st.info("No hay historial registrado aún.")
        else:
            grouped = agrupar_historial_por_mes_semana(hist_docs)
            for month_key in sorted(grouped.keys(), reverse=False):
                month_info = grouped[month_key]
                st.markdown(f"<div class='group-heading'>{month_info['label']}</div>", unsafe_allow_html=True)
                for week_no in sorted(month_info["weeks"].keys(), reverse=True):
                    week_label = "Sin semana (fecha no disponible)" if week_no == 0 else f"Semana {week_no}"
                    st.markdown(f"<div class='week-heading'>{week_label}</div>", unsafe_allow_html=True)
                    for doc_id, data, dt in month_info["weeks"][week_no]:
                        fecha_display = data.get("fecha", "-")
                        admin = data.get("admin", "Desconocido")
                        target_mx = data.get("target_mx", "-")
                        target_nombre = data.get("target_nombre", "-")
                        tipo = data.get("tipo", "-")
                        with st.container():
                            st.markdown(f"""<div class="card"><h4 style="color:#00ff99;">Alta por {admin} — {tipo}</h4>
                                <p><b>Fecha:</b> {fecha_display}</p>
                                <p><b>MX:</b> {target_mx} | <b>Nombre:</b> {target_nombre}</p>
                                </div>""", unsafe_allow_html=True)

# ---------------- FOOTER ----------------
st.markdown("""<div class="footer">DEVELOPED BY: JUAN RODRIGO GUZMÁN MARTÍNEZ</div>""", unsafe_allow_html=True)